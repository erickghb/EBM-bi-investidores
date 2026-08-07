"""
sync_all.py
-----------
Sincronização Portátil e Multi-Usuário (SharePoint + OneDrive -> Neon DB)

Esta versão utiliza caminhos dinâmicos do usuário (os.path.expanduser("~"))
para que QUALQUER máquina da EBM consiga executar o script sem nenhuma alteração no código.

Pastas Monitoradas Dinamicamente:
  - OneDrive/AUTOMATE - BI (arquivos gestao_investidores_sp.json, landbank_sp.json, investidores_sp.json)
  - OneDrive/Fluxo entrada Investidores - Pagamentos.xlsx
"""

import os
import json
import glob
import openpyxl
import psycopg2
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv

BASE_DIR = Path(__file__).resolve().parent.parent
load_dotenv(dotenv_path=BASE_DIR / ".env")

DATABASE_URL = os.getenv("DATABASE_URL")
JSON_CACHE_PATH = BASE_DIR / "api" / "fresh_fluxo_data.json"

# ── Localizador de Caminhos Dinâmicos ────────────────────────────
USER_HOME = os.path.expanduser("~")

def find_onedrive_folder(subpath):
    """Procura a pasta em múltiplos locais comuns de sincronização do OneDrive."""
    candidates = [
        os.path.join(USER_HOME, "OneDrive - EBM", subpath),
        os.path.join(USER_HOME, "OneDrive", subpath),
        os.path.join(USER_HOME, subpath),
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    return candidates[0]

def find_file_in_onedrive(filename_pattern):
    """Busca um arquivo por padrão glob em todo o diretório do OneDrive."""
    onedrive_roots = [
        os.path.join(USER_HOME, "OneDrive - EBM"),
        os.path.join(USER_HOME, "OneDrive"),
    ]
    for root in onedrive_roots:
        if os.path.exists(root):
            matches = glob.glob(os.path.join(root, "**", filename_pattern), recursive=True)
            if matches:
                return matches[0]
    return None

# ── Helper Functions ──────────────────────────────────────────────
def clean_str(val):
    if val is None:
        return None
    if isinstance(val, dict):
        val = val.get("Value") or val.get("value") or str(val)
    s = str(val).strip()
    return s if s != "" else None

def clean_num(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

def clean_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except:
            pass
    return None

def extract_field(item, *keys):
    for k in keys:
        if k in item and item[k] is not None:
            return item[k]
    return None

def sync():
    print("=" * 65)
    print("SINCRONIZACAO PORTATIL MULTI-USUARIO (ONEDRIVE + SHAREPOINT)")
    print("=" * 65)
    print(f"Usuario do Sistema: {os.getlogin()} ({USER_HOME})")

    # Localizar pastas
    automate_dir = find_onedrive_folder(os.path.join("Documentos", "AUTOMATE - BI"))
    if not os.path.exists(automate_dir):
        automate_dir = find_onedrive_folder("AUTOMATE - BI")

    print(f"Pasta do AUTOMATE - BI: {automate_dir}")

    # Localizar arquivo Excel de Fluxo e APL
    excel_path = find_file_in_onedrive("Fluxo entrada Investidores - Pagamentos.xlsx")
    if not excel_path:
        excel_path = os.path.join(
            USER_HOME, "OneDrive - EBM",
            "Intranet EBM - EIM", "BI", "Investidores",
            "Fluxo entrada Investidores - Pagamentos.xlsx"
        )
    print(f"Arquivo Excel do OneDrive: {excel_path}")

    # 1. Processar Gestão de Investidores (gestao_investidores_sp.json)
    gestao_json_path = os.path.join(automate_dir, "gestao_investidores_sp.json")
    gestao_rows = []
    if os.path.exists(gestao_json_path):
        try:
            with open(gestao_json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                items = data.get("items", data) if isinstance(data, dict) else data

            for item in items:
                gestao_id = clean_str(extract_field(item, "ID", "id"))
                cc = clean_str(extract_field(item, "Centro_x0020_de_x0020_Custo", "CentroDeCusto", "centro_custo"))
                tipo_scp = clean_str(extract_field(item, "Tipo_x0020_de_x0020_SCP", "TipoDeSCP", "tipo_scp"))
                nome_inv = clean_str(extract_field(item, "Title", "nome_investidor", "NomeInvestidor"))
                socia_ost = clean_str(extract_field(item, "S_x00f3_cia_x0020_Ostensiva", "SociaOstensiva"))
                vlr_inv = clean_num(extract_field(item, "Valor_x0020_Investido", "ValorInvestido"))
                dt_ass = clean_date(extract_field(item, "Data_x0020_assinatura_x0020_do_x", "DataAssinatura"))
                obs = clean_str(extract_field(item, "Observa_x00e7__x00f5_es", "Observacoes"))
                comissao = clean_str(extract_field(item, "Comiss_x00e3_o_x0020__x0028__x00", "Comissao"))
                status = clean_str(extract_field(item, "Status", "status"))
                ativo_inativo = clean_str(extract_field(item, "Ativo_x002f_Inativo", "AtivoInativo"))
                email = clean_str(extract_field(item, "E_x002d_mail", "Email"))
                telefones = clean_str(extract_field(item, "Telefones", "telefones"))
                status_acerto = clean_str(extract_field(item, "Status_x0020_Acerto", "StatusAcerto"))

                gestao_rows.append((
                    gestao_id, cc, cc, tipo_scp, nome_inv, socia_ost,
                    vlr_inv, dt_ass, obs, comissao, status, ativo_inativo,
                    email, telefones, status_acerto
                ))
            print(f"[1/3] SharePoint Gestao de Investidores: {len(gestao_rows)} itens lidos.")
        except Exception as e:
            print(f"[AVISO] Erro ao ler {gestao_json_path}: {e}")
    else:
        print(f"[1/3] {gestao_json_path} ainda nao gerado pelo Power Automate.")

    # 2. Processar Landbank (landbank_sp.json)
    landbank_json_path = os.path.join(automate_dir, "landbank_sp.json")
    landbank_rows = []
    if os.path.exists(landbank_json_path):
        try:
            with open(landbank_json_path, "r", encoding="utf-8") as f:
                data = json.load(f)
                items = data.get("items", data) if isinstance(data, dict) else data

            for item in items:
                titulo = clean_str(extract_field(item, "Title", "titulo", "CentroDeCusto"))
                nome = clean_str(extract_field(item, "Nome", "nome", "Imovel"))
                endereco = clean_str(extract_field(item, "Endereco", "endereco"))
                uf = clean_str(extract_field(item, "UF", "uf"))
                linha = clean_str(extract_field(item, "Linha", "linha"))
                tipologia = clean_str(extract_field(item, "Tipologia", "tipologia"))
                status = clean_str(extract_field(item, "Status", "status"))
                apl = clean_str(extract_field(item, "AplEmpreendimento", "apl_empreendimento"))

                landbank_rows.append((titulo, nome, endereco, uf, linha, tipologia, status, apl))
            print(f"[2/3] SharePoint Landbank: {len(landbank_rows)} empreendimentos lidos.")
        except Exception as e:
            print(f"[AVISO] Erro ao ler {landbank_json_path}: {e}")

    # 3. Ler Excel de Fluxo e APL
    fluxo_rows, apl_rows = [], []
    fluxo_json, apl_json = [], []

    if os.path.exists(excel_path):
        print(f"[3/3] Lendo Excel do OneDrive:\n   {excel_path}")
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

        if "FLUXO DE ENTRADA" in wb.sheetnames:
            ws_fluxo = wb["FLUXO DE ENTRADA"]
            for row in ws_fluxo.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                cc, emp, inv = clean_str(row[0]), clean_str(row[1]), clean_str(row[2])
                tit, cli = clean_str(row[3]), clean_str(row[4])
                tipo_ct, poss_mut = clean_str(row[5]), clean_str(row[6])
                status = clean_str(row[7])
                dt_pag = clean_date(row[8])
                realizado, previsto = clean_num(row[9]), clean_num(row[10])

                if not emp and not inv and not tit:
                    continue

                fluxo_rows.append((cc, emp, inv, tit, cli, tipo_ct, tipo_ct, poss_mut, status, dt_pag, realizado, previsto))
                fluxo_json.append({
                    "centro_custos": cc, "empreendimento": emp, "investidor": inv, "titulo": tit,
                    "numero_cliente": cli, "tipo_investimento": tipo_ct, "tipo_contrato": tipo_ct,
                    "possibilidade_conversao_mutuo": poss_mut, "status": status, "data_pagamento": dt_pag,
                    "realizado": realizado, "previsto": previsto
                })

        if "APL-Valor do CT" in wb.sheetnames:
            ws_apl = wb["APL-Valor do CT"]
            for row in ws_apl.iter_rows(min_row=2, values_only=True):
                if not any(row):
                    continue
                cc, emp, inv = clean_str(row[0]), clean_str(row[1]), clean_str(row[2])
                tit, cli = clean_str(row[3]), clean_str(row[4])
                tipo_ct, poss_mut = clean_str(row[5]), clean_str(row[6])
                status = clean_str(row[7])
                dt_ass = clean_date(row[8])
                area, vlr_ct = clean_num(row[9]), clean_num(row[10])

                if not emp and not inv and not tit:
                    continue

                apl_rows.append((cc, emp, inv, tit, cli, tipo_ct, poss_mut, status, dt_ass, area, vlr_ct))
                apl_json.append({
                    "centro_custos": cc, "empreendimento": emp, "investidor": inv, "titulo": tit,
                    "numero_cliente": cli, "tipo_contrato": tipo_ct, "possibilidade_conversao": poss_mut,
                    "status": status, "data_assinatura_contrato": dt_ass, "area": area, "valor_contrato": vlr_ct
                })

        wb.close()
        print(f"   -> Fluxo: {len(fluxo_rows)} parcelas | APL: {len(apl_rows)} contratos.")

    # 4. Atualizar cache JSON
    print("Atualizando arquivo de cache fresh_fluxo_data.json...")
    cache_data = {
        "fluxo": fluxo_json,
        "apl": apl_json,
        "updated_at": datetime.now().isoformat()
    }
    with open(JSON_CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(cache_data, f, ensure_ascii=False, indent=2)

    # 5. Tentar atualizar Neon Database se conexão disponível
    if DATABASE_URL:
        try:
            conn = psycopg2.connect(DATABASE_URL, connect_timeout=5)
            cur = conn.cursor()

            # Gestão de Investidores
            if gestao_rows:
                cur.execute("TRUNCATE raw.gestao_investidores RESTART IDENTITY CASCADE")
                cur.executemany("""
                    INSERT INTO raw.gestao_investidores (
                        gestao_id, obra_id, centro_custo, tipo_scp, nome_investidor, socia_ostensiva,
                        valor_investido, data_assinatura, observacoes, comissao, status, ativo_inativo,
                        email, telefones, status_acerto
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, gestao_rows)
                print(f"   -> raw.gestao_investidores atualizado ({len(gestao_rows)} registros no Neon DB)")

            # Landbank
            if landbank_rows:
                cur.execute("TRUNCATE raw.landbank RESTART IDENTITY CASCADE")
                cur.executemany("""
                    INSERT INTO raw.landbank (
                        titulo, nome, endereco, uf, linha, tipologia, status, apl_empreendimento
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                """, landbank_rows)
                print(f"   -> raw.landbank atualizado ({len(landbank_rows)} registros no Neon DB)")

            # Fluxo
            if fluxo_rows:
                cur.execute("ALTER TABLE raw.fluxo_entrada ADD COLUMN IF NOT EXISTS tipo_investimento TEXT")
                cur.execute("ALTER TABLE raw.fluxo_entrada ADD COLUMN IF NOT EXISTS tipo_contrato TEXT")
                cur.execute("ALTER TABLE raw.fluxo_entrada ADD COLUMN IF NOT EXISTS possibilidade_conversao_mutuo TEXT")
                cur.execute("TRUNCATE raw.fluxo_entrada RESTART IDENTITY CASCADE")
                cur.executemany("""
                    INSERT INTO raw.fluxo_entrada (
                        centro_custos, empreendimento, investidor, titulo, numero_cliente,
                        tipo_investimento, tipo_contrato, possibilidade_conversao_mutuo,
                        status, data_pagamento, realizado, previsto
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, fluxo_rows)

            # APL
            if apl_rows:
                cur.execute("ALTER TABLE raw.apl_valor_contrato ADD COLUMN IF NOT EXISTS valor_contrato NUMERIC")
                cur.execute("TRUNCATE raw.apl_valor_contrato RESTART IDENTITY CASCADE")
                cur.executemany("""
                    INSERT INTO raw.apl_valor_contrato (
                        centro_custos, empreendimento, investidor, titulo, numero_cliente,
                        tipo_contrato, possibilidade_conversao, status,
                        data_assinatura_contrato, area, valor_contrato
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """, apl_rows)

            conn.commit()
            cur.close()
            conn.close()
            print("[OK] Banco Neon atualizado com sucesso!")
        except Exception as e:
            print(f"[AVISO] Conexao direta ao Neon DB: {e}")

    print("\nSINCRONIZACAO COMPLETA CONCLUIDA COM SUCESSO!")
    return True

if __name__ == "__main__":
    sync()
