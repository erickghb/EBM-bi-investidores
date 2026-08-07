"""
sync_onedrive_excel.py
----------------------
Lê diretamente o arquivo Excel do OneDrive:
"C:\\Users\\erick.aires\\OneDrive - EBM\\Intranet EBM - EIM\\BI\\Investidores\\Fluxo entrada Investidores - Pagamentos.xlsx"

Carregando:
  - Aba 'FLUXO DE ENTRADA' -> raw.fluxo_entrada
  - Aba 'APL-Valor do CT'   -> raw.apl_valor_contrato

Também atualiza o arquivo api/fresh_fluxo_data.json para manter o servidor Render
sempre em sincronia perfeita!
"""

import openpyxl
import os
import json
import psycopg2
from datetime import datetime
from pathlib import Path
from dotenv import load_dotenv

BASE_DIR = Path(__file__).resolve().parent.parent
load_dotenv(dotenv_path=BASE_DIR / ".env")

DATABASE_URL = os.getenv("DATABASE_URL")
EXCEL_PATH = r"C:\Users\erick.aires\OneDrive - EBM\Intranet EBM - EIM\BI\Investidores\Fluxo entrada Investidores - Pagamentos.xlsx"
JSON_CACHE_PATH = BASE_DIR / "api" / "fresh_fluxo_data.json"

def clean_str(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s != "" else None

def clean_num(val):
    if val is None:
        return 0.0
    try:
        return float(val)
    except:
        return 0.0

def clean_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except:
            pass
    return None

def sync():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERRO: Arquivo Excel nao encontrado em: {EXCEL_PATH}")
        return False

    print(f"[1/4] Lendo arquivo Excel do OneDrive:\n   {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

    # 1. Processar FLUXO DE ENTRADA
    ws_fluxo = wb["FLUXO DE ENTRADA"]
    fluxo_rows = []
    fluxo_json = []

    for i, row in enumerate(ws_fluxo.iter_rows(min_row=2, values_only=True)):
        if not any(row):
            continue
        cc = clean_str(row[0])
        emp = clean_str(row[1])
        inv = clean_str(row[2])
        tit = clean_str(row[3])
        cli = clean_str(row[4])
        tipo_contrato = clean_str(row[5])
        poss_mutuo = clean_str(row[6])
        status = clean_str(row[7])
        dt_pag = clean_date(row[8])
        realizado = clean_num(row[9])
        previsto = clean_num(row[10])

        if not emp and not inv and not tit:
            continue

        fluxo_rows.append((
            cc, emp, inv, tit, cli,
            tipo_contrato, tipo_contrato, poss_mutuo,
            status, dt_pag, realizado, previsto
        ))

        fluxo_json.append({
            "centro_custos": cc,
            "empreendimento": emp,
            "investidor": inv,
            "titulo": tit,
            "numero_cliente": cli,
            "tipo_investimento": tipo_contrato,
            "tipo_contrato": tipo_contrato,
            "possibilidade_conversao_mutuo": poss_mutuo,
            "status": status,
            "data_pagamento": dt_pag,
            "realizado": realizado,
            "previsto": previsto
        })

    # 2. Processar APL-Valor do CT
    ws_apl = wb["APL-Valor do CT"]
    apl_rows = []
    apl_json = []

    for i, row in enumerate(ws_apl.iter_rows(min_row=2, values_only=True)):
        if not any(row):
            continue
        cc = clean_str(row[0])
        emp = clean_str(row[1])
        inv = clean_str(row[2])
        tit = clean_str(row[3])
        cli = clean_str(row[4])
        tipo_contrato = clean_str(row[5])
        poss_mutuo = clean_str(row[6])
        status = clean_str(row[7])
        dt_ass = clean_date(row[8])
        area = clean_num(row[9])
        vlr_ct = clean_num(row[10])

        if not emp and not inv and not tit:
            continue

        apl_rows.append((
            cc, emp, inv, tit, cli,
            tipo_contrato, poss_mutuo, status,
            dt_ass, area, vlr_ct
        ))

        apl_json.append({
            "centro_custos": cc,
            "empreendimento": emp,
            "investidor": inv,
            "titulo": tit,
            "numero_cliente": cli,
            "tipo_contrato": tipo_contrato,
            "possibilidade_conversao": poss_mutuo,
            "status": status,
            "data_assinatura_contrato": dt_ass,
            "area": area,
            "valor_contrato": vlr_ct
        })

    wb.close()

    print(f"[2/4] Excel lido com sucesso: {len(fluxo_rows)} linhas de Fluxo, {len(apl_rows)} linhas de APL.")

    # 3. Atualizar fresh_fluxo_data.json
    print(f"[3/4] Atualizando arquivo de cache {JSON_CACHE_PATH.name}...")
    cache_data = {
        "fluxo": fluxo_json,
        "apl": apl_json,
        "updated_at": datetime.now().isoformat()
    }
    with open(JSON_CACHE_PATH, "w", encoding="utf-8") as f:
        json.dump(cache_data, f, ensure_ascii=False, indent=2)

    # 4. Tentar atualizar Neon Database se conexao disponivel
    if DATABASE_URL:
        print("[4/4] Conectando ao banco Neon PostgreSQL...")
        try:
            conn = psycopg2.connect(DATABASE_URL, connect_timeout=5)
            cur = conn.cursor()

            cur.execute("ALTER TABLE raw.fluxo_entrada ADD COLUMN IF NOT EXISTS tipo_investimento TEXT")
            cur.execute("ALTER TABLE raw.fluxo_entrada ADD COLUMN IF NOT EXISTS tipo_contrato TEXT")
            cur.execute("ALTER TABLE raw.fluxo_entrada ADD COLUMN IF NOT EXISTS possibilidade_conversao_mutuo TEXT")
            cur.execute("TRUNCATE raw.fluxo_entrada RESTART IDENTITY CASCADE")

            cur.executemany("""
                INSERT INTO raw.fluxo_entrada (
                    centro_custos, empreendimento, investidor, titulo, numero_cliente,
                    tipo_investimento, tipo_contrato, possibilidade_conversao_mutuo,
                    status, data_pagamento, realizado, previsto
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, fluxo_rows)

            cur.execute("ALTER TABLE raw.apl_valor_contrato ADD COLUMN IF NOT EXISTS valor_contrato NUMERIC")
            cur.execute("TRUNCATE raw.apl_valor_contrato RESTART IDENTITY CASCADE")

            cur.executemany("""
                INSERT INTO raw.apl_valor_contrato (
                    centro_custos, empreendimento, investidor, titulo, numero_cliente,
                    tipo_contrato, possibilidade_conversao, status,
                    data_assinatura_contrato, area, valor_contrato
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, apl_rows)

            conn.commit()
            cur.close()
            conn.close()
            print("[4/4] Banco Neon atualizado com sucesso!")
        except Exception as e:
            print(f"[AVISO] Conexao direta ao Neon falhou ({e}). Cache local atualizado em fresh_fluxo_data.json.")

    print("[SUCESSO] Sincronizacao concluida com sucesso!")
    return True

if __name__ == "__main__":
    sync()
