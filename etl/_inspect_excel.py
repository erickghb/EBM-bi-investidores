import openpyxl
path = r"C:\Users\erick.aires\OneDrive - EBM\Intranet EBM - EIM\BI\Investidores\Fluxo entrada Investidores - Pagamentos.xlsx"
wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    headers = [c.value for c in next(ws.iter_rows(max_row=1))]
    print(f"=== SHEET: {sheet_name} ===")
    print("Headers:", headers)
    for i, row in enumerate(ws.iter_rows(min_row=2, max_row=3, values_only=True)):
        print(f"Row {i+2}:", row)
wb.close()
